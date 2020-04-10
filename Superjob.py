import time
import selenium
from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

################### Xpath Для специализации ################################



################### Xpath Для подрубрик ################################
xpath_rubriks = {
    'IT, Интернет, связь, телеком': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[1]/div/div/button',
    'Административная работа, секретариат, АХО': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[2]/div/div/button',

    'Банки, инвестиции, лизинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[3]/div/div/button',
    'Безопасность, службы охраны': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[4]/div/div/button',
    'Бухгалтерия, финансы, аудит': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[5]/div/div/button',

    'Дизайн': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[6]/div/div/button',
    'Домашний персонал': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[7]/div/div/button',
    'Закупки, снабжение': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[8]/div/div/button',
    'Искусство, культура, развлечения': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[9]/div/div/button',
    'Кадры, управление персоналом': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[10]/div/div/button',
    'Консалтинг, стратегическое развитие': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[11]/div/div/button',
    'Маркетинг, реклама, PR': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[12]/div/div/button',
    'Медицина, фармацевтика, ветеринария': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[13]/div/div/button',

    'Наука, образование, повышение квалификации': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[14]/div/div/button',
    'Некоммерческие организации, волонтерство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[15]/div/div/button',
    'Продажи': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[16]/div/div/button',

    'Промышленность, производство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[17]/div/div/button',
    'Рабочий персонал': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[18]/div/div/button',

    'СМИ, издательства': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[19]/div/div/button',
    'Сельское хозяйство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[20]/div/div/button',

    'Спорт, фитнес, салоны красоты, SPA': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[21]/div/div/button',
    'Страхование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[22]/div/div/button',

    'Строительство, проектирование, недвижимость': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[23]/div/div/button',
    'Сырье': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[24]/div/div/button',

    'Топ-персонал': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[25]/div/div/button',
    'Транспорт, логистика, ВЭД': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[26]/div/div/button',
    'Туризм, гостиницы, общественное питание': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[27]/div/div/button',
    'Услуги, ремонт, сервисное обслуживание': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[28]/div/div/button',
    'Юриспруденция': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[1]/ul/li[29]/div/div/button',
}

################### Xpath Для подрубрик ################################

# IT / Интернет / Телеком
xpath_it_ithernet = {

    'CRM-системы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Web, UI, UX дизайн': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Web-верстка': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Администрирование баз данных': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    #
    'Аналитика': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Внедрение и сопровождение ПО': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Защита информации': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Игровое ПО / Геймдевелопмент': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    #
    'Инжиниринг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Интернет, создание и поддержка сайтов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Киберспорт': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    #
    'Компьютерная анимация и мультимедиа': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Контент': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    #
    'Мобильная разработка': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    'Оптимизация, SEO': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[15]/div/div/label/div/div[1]/span',
    'Передача данных и доступ в интернет': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[16]/div/div/label/div/div[1]/span',
    'Разработка и сопровождение банковского ПО': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[17]/div/div/label/div/div[1]/span',
    'Разработка, программирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[18]/div/div/label/div/div[1]/span',
    #
    'Сетевые технологии': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[19]/div/div/label/div/div[1]/span',
    'Системная интеграция': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[20]/div/div/label/div/div[1]/span',
    'Системное администрирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[21]/div/div/label/div/div[1]/span',
    #
    'Системы автоматизированного проектирования (САПР)': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[22]/div/div/label/div/div[1]/span',
    'Системы управления предприятием (ERP)': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[23]/div/div/label/div/div[1]/span',
    'Сотовые, беспроводные технологии': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[24]/div/div/label/div/div[1]/span',
    #
    'Телекоммуникации и связь': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[25]/div/div/label/div/div[1]/span',
    'Тестирование, QA': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[26]/div/div/label/div/div[1]/span',
    'Техническая документация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[27]/div/div/label/div/div[1]/span',
    'Техническая поддержка': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[28]/div/div/label/div/div[1]/span',
    #
    'Управление продуктом': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[29]/div/div/label/div/div[1]/span',
    'Управление проектами': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[30]/div/div/label/div/div[1]/span',
    #
    'Электронная коммерция': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[31]/div/div/label/div/div[1]/span',
    'Электронный документооборот': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[32]/div/div/label/div/div[1]/span',
    'Юзабилити': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[33]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[34]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[35]/div/div/label/div/div[1]/span',
}

# Административная работа, секретариат, АХО
xpath_it_adm_personal = {

    'Архивное дело': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'АХО': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Делопроизводство, ввод данных, систематизация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Диспетчерская служба': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    #
    'Курьерская служба': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Переводы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Секретариат, ресепшн, офис-менеджмент': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
}

# Транспорт, логистика, ВЭД
xpath_logistika = {
    'Авиаперевозки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Автоперевозки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[2]/span',
    'ВЭД': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',

    'Железнодорожные перевозки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Контейнерные перевозки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',

    'Логистика': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Метрополитен': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',

    'Морские, речные перевозки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Складское хозяйство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Таможня': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Трубопроводы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',

    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
}

# Банки, инвестиции, лизинг
xpath_banki_invest = {

    'Банковская бухгалтерия': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Бэк-Офис': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Бюджетирование и планирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Валютные операции': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    #
    'Вклады': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Депозитарий': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Документарные операции': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Залоги и проблемная задолженность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    #
    'Ипотека': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Кредитование физических лиц': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Кредитование юридических лиц': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    #
    'Лизинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Методология, разработка и продажа корпоративных продуктов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    #
    'Методология, разработка и продажа продуктов Private Banking': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    'Методология, разработка и продажа розничных продуктов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[15]/div/div/label/div/div[1]/span',
    'Налоговый учёт': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[16]/div/div/label/div/div[1]/span',
    'Обслуживание банкоматов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[17]/div/div/label/div/div[1]/span',
    'Пластиковые карты (эквайринг)': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[18]/div/div/label/div/div[1]/span',
    #
    'Продажа банковских продуктов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[19]/div/div/label/div/div[1]/span',
    'Разработка банковских продуктов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[20]/div/div/label/div/div[1]/span',
    'Расчёты и обработка платежей, касса': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[21]/div/div/label/div/div[1]/span',
    #
    'Риски': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[22]/div/div/label/div/div[1]/span',
    'Торговое финансирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[23]/div/div/label/div/div[1]/span',
    'Управление активами': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[24]/div/div/label/div/div[1]/span',
    #
    'Управление ликвидностью и балансовыми рисками': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[25]/div/div/label/div/div[1]/span',
    'Управленческая отчетность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[26]/div/div/label/div/div[1]/span',
    'Финансовая отчётность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[27]/div/div/label/div/div[1]/span',
    'Финансовый анализ и контроль': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[28]/div/div/label/div/div[1]/span',
    #
    'Ценные бумаги': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[29]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[30]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[31]/div/div/label/div/div[1]/span',
}

# Безопасность, службы охраны
xpath_ohrana= {
    'Видеонаблюдение': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Имущественная безопасность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Инкассация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',

    'Информационная безопасность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Кинология': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',

    'Личная безопасность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Охранно-, детективная деятельность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',

    'Охранные службы предприятий': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Пожарная безопасность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Служба спасения': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'ЧОП': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',

    'Экономическая безопасность': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    }

# Бухгалтерия, финансы, аудит
xparh_buhgalteria = {

    'GAAP': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Аудит, ревизия, инспектирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Бухгалтерский учет': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Бюджетирование и планирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Казначейство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Кассовые операции': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Кредитный контроль': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'МСФО, IFRS, РСБУ': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Налоги': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Основные средства': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Первичная документация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Планово-экономическое управление': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Расчет себестоимости': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    'Учет заработной платы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    'Учет счетов и платежей': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[15]/div/div/label/div/div[1]/span',
    'Учет труда и социальных выплат сотрудникам': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[16]/div/div/label/div/div[1]/span',
    'Учет фондов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[17]/div/div/label/div/div[1]/span',
    'Финансовый анализ и контроль': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[18]/div/div/label/div/div[1]/span',
    'Финансовый менеджмент': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[19]/div/div/label/div/div[1]/span',
    'Экономический анализ': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[20]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[21]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[22]/div/div/label/div/div[1]/span',
}

# Дизайн
xparh_design = {

    'Архитектура': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Аудио, видео': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Верстка, дизайн': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Графическое искусство, иллюстрации': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Дизайн интерьера': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Компьютерная анимация и мультимедиа': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Ландшафтный дизайн': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Мода и Аксессуары': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Промышленный дизайн': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Рекламный дизайн': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Фотография': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
}

# Домашний персонал
xparh_home_personal = {

    'Ведение хозяйства': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Домашний уход за больными, престарелыми, инвалидами': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Домашний уход за детьми': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Уборка и помощь по дому': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
}

# Закупки, снабжение
xparh_zakupki = {

    'Автомобили, запчасти': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Алкоголь, напитки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Бытовая техника, электроника, фото, видео': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Вендинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Госзакупки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'ГСМ, нефть': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Зоотовары, ветеринарные препараты': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Канцелярские товары': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Книги, печатные издания': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Компьютеры, оргтехника, ПО': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Косметика, бытовая химия': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Мебель': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Медицина, фармацевтика': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    'Металлопрокат': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    'Оборудование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[15]/div/div/label/div/div[1]/span',
    'Продукты питания': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[16]/div/div/label/div/div[1]/span',
    'Продукция химического производства': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[17]/div/div/label/div/div[1]/span',
    'Сантехника': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[18]/div/div/label/div/div[1]/span',
    'Сельское хозяйство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[19]/div/div/label/div/div[1]/span',
    'Системы безопасности': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[20]/div/div/label/div/div[1]/span',
    'Спортивные товары и фитнес-услуги': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[21]/div/div/label/div/div[1]/span',
    'Строительно-отделочные материалы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[22]/div/div/label/div/div[1]/span',
    'Сырье': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[23]/div/div/label/div/div[1]/span',
    'Табачная продукция': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[24]/div/div/label/div/div[1]/span',
    'Текстиль, одежда, обувь': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[25]/div/div/label/div/div[1]/span',
    'Телекоммуникации, сетевые решения, средства связи': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[26]/div/div/label/div/div[1]/span',
    'Тендеры': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[27]/div/div/label/div/div[1]/span',
    'Товары народного потребления': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[28]/div/div/label/div/div[1]/span',
    'Услуги для бизнеса': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[29]/div/div/label/div/div[1]/span',
    'Флористика': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[30]/div/div/label/div/div[1]/span',
    'Электротехническое оборудование, светотехника': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[32]/div/div/label/div/div[1]/span',
    'Ювелирные изделия': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[33]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[34]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[35]/div/div/label/div/div[1]/span',
}

# Искусство, культура, развлечения
xparh_kultura = {

    'Балет, хореография, танцы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Библиотеки': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Декоративно-прикладное искусство, народные промыслы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Живопись, графика, скульптура': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Искусствоведение': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Кино, мультипликация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Литература': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Модельный бизнес': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Музей, выставочный зал': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Музыка, пение': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Развлечения, игры, отдых, анимация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Реставрация, антиквариат': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Театр': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    'Шоу-бизнес': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[15]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[16]/div/div/label/div/div[1]/span',
}

# Кадры, управление персоналом
xparh_hr_kadri = {

    'Кадровый учет, делопроизводство, кадровое администрирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Оплата труда, компенсации  и льготы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Охрана труда': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Психология труда и социальная психология': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Развитие персонала, обучение, тренинги': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Рекрутмент, подбор персонала': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Управление персоналом': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
}

# Консалтинг, стратегическое развитие
xparh_konsalting = {

    'IT-консалтинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Кадровый консалтинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Корпоративные финансы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Маркетинговый и PR консалтинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Реинжиниринг бизнес-процессов': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Стратегический консалтинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Управление проектами': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Управленческий консалтинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Экологический консалтинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
}

# Маркетинг, реклама, PR
xparh_marketing = {

    'Маркетинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Медиаисследования, рейтинги': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Медиапланирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Мерчандайзинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Наружная реклама': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Печатная реклама': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Политический PR': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Продвижение, специальные мероприятия': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Производство рекламы': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Радиореклама': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Рекламно-сувенирная продукция': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Рекламное агентство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Социологические исследования': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    'Телевизионная реклама': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
    'Телемаркетинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[15]/div/div/label/div/div[1]/span',
    'Торговый маркетинг': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[16]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[17]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[18]/div/div/label/div/div[1]/span',
}

# Медицина, фармацевтика, ветеринария
xparh_medicina = {

    'Больницы, поликлиники, диагностические центры, лаборатории': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Ветеринария': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Клинические испытания': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Лечебное дело (врачи-специалисты)': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Лицензирование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Медицинское оборудование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Санитарно-эпидемиологический надзор': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'Сертификация': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Средний и младший медицинский персонал': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Товары медицинского назначения': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Уход за больными': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Фармацевтика': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[14]/div/div/label/div/div[1]/span',
}

# Наука, образование, повышение квалификации
xparh_obrazovanie = {

    'Академия наук': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[1]/div/div/label/div/div[1]/span',
    'Бизнес-образование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[2]/div/div/label/div/div[1]/span',
    'Внешкольное образование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[3]/div/div/label/div/div[1]/span',
    'Высшее образование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[4]/div/div/label/div/div[1]/span',
    'Дошкольное воспитание и образование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[5]/div/div/label/div/div[1]/span',
    'Инновационные технологии': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[6]/div/div/label/div/div[1]/span',
    'Курсы, тренинги, семинары, повышение квалификации': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[7]/div/div/label/div/div[1]/span',
    'НИИ, КБ': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[8]/div/div/label/div/div[1]/span',
    'Репетиторство': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[9]/div/div/label/div/div[1]/span',
    'Среднее образование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[10]/div/div/label/div/div[1]/span',
    'Среднее специальное образование': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[11]/div/div/label/div/div[1]/span',
    'Другое': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[12]/div/div/label/div/div[1]/span',
    'Начало карьеры, мало опыта': '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul/li[13]/div/div/label/div/div[1]/span',
}

vibor_podrubrik = {
    'IT, Интернет, связь, телеком' : xpath_it_ithernet,
    'Административная работа, секретариат, АХО' : xpath_it_adm_personal,
    'Банки, инвестиции, лизинг' : xpath_banki_invest,
    'Транспорт, логистика, ВЭД' : xpath_logistika,
    'Безопасность, службы охраны' : xpath_ohrana,
    'Бухгалтерия, финансы, аудит' : xparh_buhgalteria,
    'Дизайн' : xparh_design,
    'Домашний персонал' : xparh_home_personal,
    'Закупки, снабжение' : xparh_zakupki,
    'Исскуство, культура, развлечения' : xparh_kultura,
    'Кадры, управление персоналом' : xparh_hr_kadri,
    'Консалтинг, стратегическое развитие' : xparh_konsalting,
    'Маркетинг, реклама, PR' : xparh_marketing,
    'Медицина,  фармацевтика, ветеринария' : xparh_medicina,
    'Наука, образование, повышение квалификации' : xparh_obrazovanie

    }

################### считываем данные с excel ################################

wb = xlrd.open_workbook('./DATA.xlsx')
sheet_xlr = wb.sheet_by_name('Вакансии')
a = 2


rows_sheet = sheet_xlr.nrows
# print(rows_sheet)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)

sheet = wb[lst[0]]
sheet.title
error_string = 1056

# random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.

error_string = 1060
i = random.randint(3, rows_sheet)
################### заводим содержимое в переменные ################################
# print(a)
vacancy = (sheet.cell(row=i, column=1).value)
zp_ot = (sheet.cell(row=i, column=2).value)
zp_do = (sheet.cell(row=i, column=3).value)
opyt = (sheet.cell(row=i, column=4).value)
opisanie = (sheet.cell(row=i, column=5).value)
grafik = (sheet.cell(row=i, column=6).value)
company = (sheet.cell(row=i, column=7).value)
opis_company = (sheet.cell(row=i, column=8).value)
company1 = (sheet.cell(row=i, column=9).value)
opis_company1 = (sheet.cell(row=i, column=10).value)
company2 = (sheet.cell(row=i, column=11).value)
opis_company2 = (sheet.cell(row=i, column=12).value)
company3 = (sheet.cell(row=i, column=13).value)
opis_company3 = (sheet.cell(row=i, column=14).value)
company4 = (sheet.cell(row=i, column=15).value)
opis_company4 = (sheet.cell(row=i, column=16).value)
sfera_deyatel = (sheet.cell(row=i, column=17).value)
opis_deyatel = (sheet.cell(row=i, column=18).value)
name_company = (sheet.cell(row=i, column=19).value)
opis_deyatel_rename_company = (sheet.cell(row=i, column=20).value)

# print(vacancy)
error_string = 1099
# СЧИТЫВАЕМ ГОРОДА иЗ ЭКСЕЛЯ ДИНАМИЧЕСКИ

sheet = wb[lst[1]]
sheet.title

barnaul_adr = []
volgograd_adr = []
voronej_adr = []
ekaterenburg_adr = []
izevsk_adr = []
irkuts_adr = []
kazan_adr = []
kaliningrad_adr = []
kemerovo_adr = []
krasnodar_adr = []
krasnoyars_adr = []
moskow_adr = []
naber_cheln_adr = []
nizniy_novgorod_adr = []
novosib_adr = []
omsk_adr = []
orenburg_adr = []
perm_adr = []
rostov_na_dony_adr = []
ryazan_adr = []
samara_adr = []
sankt_peter_adr = []
saratov_adr = []
sochi_and_adler = []
toliatty_adr = []
tomsk_adr = []
tyla_adr = []
tumen_adr = []
ufa_adr = []
chelabinsk_adr = []
yaroslavl_adr = []
musor_exela = []
#goroda_arr = [musor_exela, barnaul_adr, volgograd_adr, voronej_adr, ekaterenburg_adr, izevsk_adr, irkuts_adr, kazan_adr, kaliningrad_adr, kemerovo_adr, krasnodar_adr, krasnoyars_adr, moskow_adr, naber_cheln_adr, nizniy_novgorod_adr, novosib_adr, omsk_adr, orenburg_adr, perm_adr, rostov_na_dony_adr, ryazan_adr, samara_adr, sankt_peter_adr, saratov_adr, sochi_and_adler, toliatty_adr, tomsk_adr, tyla_adr, tumen_adr, ufa_adr, chelabinsk_adr, yaroslavl_adr]
goroda_arr = []

kolichestvo_gorodov = 0

for cell in sheet['A']:
    kolichestvo_gorodov += 1
    print(cell.value)

for number_gorod in range(kolichestvo_gorodov):
    goroda_arr.append([])
    for cell in list(sheet.rows)[number_gorod]:
         if cell.value != None:
            goroda_arr[number_gorod].append(cell.value)

kolichestvo_adresov = []
high_number_gorod = 0
number_goroda = 0
while number_goroda < len(goroda_arr):
    max_number = len(goroda_arr[number_goroda])
    kolichestvo_adresov.append(len(goroda_arr[number_goroda]))
    if high_number_gorod< max_number:
        high_number_gorod = max_number
        number_goroda = number_goroda + 1
    else:
        number_goroda = number_goroda + 1

###############    autorith  ###############

n = 0
while n < 5:
    sheet = wb[lst[2]]
    sheet.title
    login = (sheet.cell(row=1, column=2).value)
    password = (sheet.cell(row=2, column=2).value)

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    browser = webdriver.Chrome(chrome_options=options)
    driver = browser

    browser.get('https://nn.superjob.ru/hr/vacancy/create/')



    ############### Login ########################

    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[3]/div/div/div/div/div[1]/label/div/div/input').send_keys(login)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[3]/div/div/div/div/div[2]/label/div/div[1]/input').send_keys(password)
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[4]/div/div[1]/button').click()

    ############### Должность ####################
    time.sleep(4)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[1]/div/div[2]/div/div[1]/label/div/div/input').send_keys(vacancy)

    ############## Специализация #################
    companys = [company, company1, company2, company3, company4]
    opis_companys = [opis_company, opis_company1, opis_company2, opis_company3, opis_company4]
    iter_companys = 0 #Для теста пока без цикла, потом добавить как в работару
    print(companys)
    print(opis_companys)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[2]/div/div/div/button').click()
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[3]/div/button/span/span/span').click()
    except:
        print('чисто')
    time.sleep(1)
    while iter_companys < 5:
        if iter_companys != 0:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[2]/div/div/div/button').click()
        rubrika = xpath_rubriks.get(companys[iter_companys])
        browser.find_element_by_xpath(rubrika).click()
        time.sleep(1)
        xpartt = vibor_podrubrik.get(companys[iter_companys])
        xpartt2 = xpartt.get(opis_companys[iter_companys])
        browser.find_element_by_xpath(xpartt2).click()
        time.sleep(1)
        browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[3]/div/div/div[1]/button/span/span/span').click()
        time.sleep(1)

        iter_companys+=1

    ############# Занятость #################### НУЖНО СДЕЛАТЬ

    browser.find_element_by_xpath('//*[@id="detailInfo.workType.id-input"]').click()
    browser.find_element_by_xpath('//*[@id="detailInfo.workType.id-input"]').send_keys(grafik + Keys.ENTER)
    ############# Уровень дохода ###############

    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[7]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[1]/label/div/div[2]/input').send_keys(zp_ot)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[7]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[2]/label/div/div[2]/input').send_keys(zp_do)

    ############ Опыт работы ###################
    try:
        if opyt == 'не имеет значения':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[1]').click()
        elif opyt == '1 год':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[3]').click()
        elif opyt == '3 года':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[4]').click()
        elif opyt == '6 лет':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[5]').click()
        elif opyt == 'без опыта':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[2]').click()
    except:
        print('Выбраный опыт работы уже стоит')

    ############ Описание вакансии #############

    #Вытаскиваем Условия
    frag_opisanie = opisanie.partition('Условия:')
    frag_opisanie = frag_opisanie[2].partition('Обязанности:')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    yslovia = frag_opisanie[0]

    #Вытаскиваем Обязанности
    frag_opisanie = opisanie.partition('Обязанности:')
    frag_opisanie = frag_opisanie[2].partition('Требования:')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    obiaznosti = frag_opisanie[0]

    #Вытаскиваем Требования
    frag_opisanie = opisanie.partition('Требования:')
    frag_opisanie = frag_opisanie[2].partition('\n \n')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    trebovania = frag_opisanie[0]

    #Вытаскиваем шапку
    frag_opisanie = opisanie.partition('Условия:')
    shapka = frag_opisanie[0]
    print(frag_opisanie)

    #Вытаскиваем подвал
    frag_opisanie = opisanie.partition('\n \n')
    podval = frag_opisanie[2]

    #Заполняем описание
    linii = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[10]/div/div/div[2]/div/div[1]/div/div/div/div[2]')
    linii.find_element_by_tag_name('b').send_keys(Keys.ARROW_UP + Keys.ENTER)
    lst_linii = linii.find_elements_by_tag_name('li')
    linii.find_element_by_tag_name('p').send_keys(Keys.LEFT_CONTROL + 'b' + Keys.LEFT_CONTROL + (shapka + Keys.BACKSPACE))
    lst_linii[0].send_keys(obiaznosti + Keys.BACKSPACE + Keys.DELETE)
    lst_linii[2].send_keys(trebovania + Keys.DELETE)
    lst_linii[4].send_keys(yslovia + Keys.BACKSPACE + Keys.DELETE)
    linii.send_keys('\n\n' + Keys.LEFT_CONTROL + 'b' + Keys.LEFT_CONTROL + podval)



    #Заполняем города
    #ля теста iter_gorod
    iter_gorod = 1
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[1]/div/div/div[2]/ul/li[1]/span/span[2]/button').click()
    except:
        print('город автоматически не проставился')

    adr_bar = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[1]/div/div/div[3]/div/div[1]/label/div/div/input')
    random_adres = random.randint(1, kolichestvo_adresov[iter_gorod]-1)
    print('random adres = ', random_adres)
    adr_bar.send_keys(goroda_arr[iter_gorod][0])

    adr_bar = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[2]/div/div/div/div[2]/div/div[1]/div/div[1]/label/div/div/input')
    adr_bar.send_keys(goroda_arr[iter_gorod][random_adres])
    time.sleep(2)
    adr_bar.send_keys(Keys.ARROW_DOWN + Keys.ENTER)

    ############## Информация о компании ###################################

    if sfera_deyatel != None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/label/div/div[1]/span').click()
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[2]/button').click()
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[1]/div/label/div/div/input').send_keys(sfera_deyatel)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[2]/div/label/div/div/textarea').send_keys(opis_deyatel)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[3]/div/div[1]/button').click()
        except:
            print("Скрыть информацию не получилось")
    elif name_company != None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[2]/button').click()
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[1]/div/label/div/div/input').send_keys(name_company)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[2]/div/label/div/div/textarea').send_keys(opis_deyatel_rename_company)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[3]/div/div[1]/button').click()
        except:
            print("Отредактировать информацию не получилось")
    else:
        print("Оставляем название компании ткаим какое оно есть")


    ############## Размещаем вакансию и закрываем браузер ##################
    time.sleep(1000)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[6]/div/div[2]/div/div[1]/div/div[1]/button/span/span/span').click()
    time.sleep(1)
    browser.quit()
    n+=1


#browser.quit()