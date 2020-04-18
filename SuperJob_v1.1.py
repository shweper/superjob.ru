import time
import selenium
from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

################### Xpath Для специализации ################################



################### Xpath Для подрубрик ################################

xpath_specializacii = {
    'IT, Интернет, связь, телеком': '0',
    'Административная работа, секретариат, АХО': '1',
    'Банки, инвестиции, лизинг': '2',
    'Безопасность, службы охраны': '3',
    'Бухгалтерия, финансы, аудит': '4',
    'Дизайн': '5',
    'Домашний персонал': '6',
    'Закупки, снабжение': '7',
    'Искусство, культура, развлечения': '8',
    'Кадры, управление персоналом': '9',
    'Консалтинг, стратегическое развитие': '10',
    'Маркетинг, реклама, PR': '11',
    'Медицина, фармацевтика, ветеринария': '12',
    'Наука, образование, повышение квалификации': '13',
    'Некоммерческие организации, волонтерство': '14',
    'Продажи': '15',
    'Промышленность, производство': '16',
    'Рабочий персонал': '17',
    'СМИ, издательства': '18',
    'Сельское хозяйство': '19',
    'Спорт, фитнес, салоны красоты, SPA': '20',
    'Страхование': '21',
    'Строительство, проектирование, недвижимость': '22',
    'Сырье': '23',
    'Топ-персонал': '24',
    'Транспорт, логистика, ВЭД': '25',
    'Туризм, гостиницы, общественное питание': '26',
    'Услуги, ремонт, сервисное обслуживание': '27',
    'Юриспруденция': '28'
}

################### Xpath Для подрубрик ################################

# IT / Интернет / Телеком
xpath_it_ithernet = {

'CRM-системы':'1',
'Web, UI, UX дизайн':'3',
'Web-верстка':'5',
'Администрирование баз данных':'7',
'Аналитика':'9',
'Внедрение и сопровождение ПО':'11',
'Защита информации':'13',
'Игровое ПО / Геймдевелопмент':'15',
'Инжиниринг':'17',
'Интернет, создание и поддержка сайтов':'19',
'Киберспорт':'21',
'Компьютерная анимация и мультимедиа':'23',
'Контент':'25',
'Мобильная разработка':'27',
'Оптимизация, SEO':'29',
'Передача данных и доступ в интернет':'31',
'Разработка и сопровождение банковского ПО':'33',
'Разработка, программирование':'35',
'Сетевые технологии':'37',
'Системная интеграция':'39',
'Системное администрирование':'41',
'Системы автоматизированного проектирования (САПР)':'43',
'Системы управления предприятием (ERP)':'45',
'Сотовые, беспроводные технологии':'47',
'Телекоммуникации и связь':'49',
'Тестирование, QA':'51',
'Техническая документация':'53',
'Техническая поддержка':'55',
'Управление продуктом':'57',
'Управление проектами':'59',
'Электронная коммерция':'61',
'Электронный документооборот':'63',
'Юзабилити':'65',
'Другое':'67',
'Начало карьеры, мало опыта':'69',

}

# Административная работа, секретариат, АХО
xpath_it_adm_personal = {

'Архивное дело':'1',
'АХО':'3',
'Делопроизводство, ввод данных, систематизация':'5',
'Диспетчерская служба':'7',
'Курьерская служба':'9',
'Переводы':'11',
'Секретариат, ресепшн, офис-менеджмент':'13',
'Другое':'15',
'Начало карьеры, мало опыта':'17',

}

# Транспорт, логистика, ВЭД
xpath_logistika = {

'Авиаперевозки':'1',
'Автоперевозки':'3',
'ВЭД':'5',
'Железнодорожные перевозки':'7',
'Контейнерные перевозки':'9',
'Логистика':'11',
'Метрополитен':'13',
'Морские, речные перевозки':'15',
'Складское хозяйство':'17',
'Таможня':'19',
'Трубопроводы':'21',
'Другое':'23',
'Начало карьеры, мало опыта':'25',
}

# Банки, инвестиции, лизинг
xpath_banki_invest = {

'Банковская бухгалтерия':'1',
'Бэк-Офис':'3',
'Бюджетирование и планирование':'5',
'Валютные операции':'7',
'Вклады':'9',
'Депозитарий':'11',
'Документарные операции':'13',
'Залоги и проблемная задолженность':'15',
'Ипотека':'17',
'Кредитование физических лиц':'19',
'Кредитование юридических лиц':'21',
'Лизинг':'23',
'Методология, разработка и продажа корпоративных продуктов':'25',
'Методология, разработка и продажа продуктов Private Banking':'27',
'Методология, разработка и продажа розничных продуктов':'29',
'Налоговый учёт':'31',
'Обслуживание банкоматов':'33',
'Операции с драгметаллами':'35',
'Пластиковые карты (эквайринг)':'37',
'Продажа банковских продуктов':'39',
'Разработка банковских продуктов':'41',
'Расчёты и обработка платежей, касса':'43',
'Риски':'45',
'Торговое финансирование':'47',
'Управление активами':'49',
'Управление ликвидностью и балансовыми рисками':'51',
'Управленческая отчетность':'53',
'Финансовая отчётность':'55',
'Финансовый анализ и контроль':'57',
'Ценные бумаги':'59',
'Другое':'61',
'Начало карьеры, мало опыта':'63',
}

# Безопасность, службы охраны
xpath_ohrana= {

'Видеонаблюдение':'1',
'Имущественная безопасность':'3',
'Инкассация':'5',
'Информационная безопасность':'7',
'Кинология':'9',
'Личная безопасность':'11',
'Охранно-, детективная деятельность':'13',
'Охранные службы предприятий':'15',
'Пожарная безопасность':'17',
'Служба спасения':'19',
'ЧОП':'21',
'Экономическая безопасность':'23',
'Другое':'25',
'Начало карьеры, мало опыта':'27',
    }

# Бухгалтерия, финансы, аудит
xparh_buhgalteria = {

'GAAP':'1',
'Аудит, ревизия, инспектирование':'3',
'Бухгалтерский учет':'5',
'Бюджетирование и планирование':'7',
'Казначейство':'9',
'Кассовые операции':'11',
'Кредитный контроль':'13',
'МСФО, IFRS, РСБУ':'15',
'Налоги':'17',
'Основные средства':'19',
'Первичная документация':'21',
'Планово-экономическое управление':'23',
'Расчет себестоимости':'25',
'Учет заработной платы':'27',
'Учет счетов и платежей':'29',
'Учет труда и социальных выплат сотрудникам':'31',
'Учет фондов':'33',
'Финансовый анализ и контроль':'35',
'Финансовый менеджмент':'37',
'Экономический анализ':'39',
'Другое':'41',
'Начало карьеры, мало опыта':'43',
}

# Дизайн
xparh_design = {

    'Архитектура': '1',
    'Аудио, видео': '3',
    'Верстка, дизайн': '5',
    'Графическое искусство, иллюстрации': '7',
    'Дизайн интерьера': '9',
    'Компьютерная анимация и мультимедиа': '11',
    'Ландшафтный дизайн': '13',
    'Мода и Аксессуары': '15',
    'Промышленный дизайн': '17',
    'Рекламный дизайн': '19',
    'Фотография': '21',
    'Другое': '23',
    'Начало карьеры, мало опыта': '25',

}

# Домашний персонал
xparh_home_personal = {

'Ведение хозяйства':'1',
'Домашний уход за больными, престарелыми, инвалидами':'3',
'Домашний уход за детьми':'5',
'Уборка и помощь по дому':'7',
'Другое':'9',
'Начало карьеры, мало опыта':'11',

}

# Закупки, снабжение
xparh_zakupki = {

    'Автомобили, запчасти': '1',
    'Алкоголь, напитки': '3',
    'Бытовая техника, электроника, фото, видео': '5',
    'Вендинг': '7',
    'Госзакупки': '9',
    'ГСМ, нефть': '11',
    'Зоотовары, ветеринарные препараты': '13',
    'Канцелярские товары': '15',
    'Книги, печатные издания': '17',
    'Компьютеры, оргтехника, ПО': '19',
    'Косметика, бытовая химия': '21',
    'Мебель': '23',
    'Медицина, фармацевтика': '25',
    'Металлопрокат': '27',
    'Оборудование': '29',
    'Продукты питания': '31',
    'Продукция химического производства': '33',
    'Сантехника': '35',
    'Сельское хозяйство': '37',
    'Системы безопасности': '39',
    'Спортивные товары и фитнес-услуги': '41',
    'Строительно-отделочные материалы': '43',
    'Сырье': '45',
    'Табачная продукция': '47',
    'Текстиль, одежда, обувь': '49',
    'Телекоммуникации, сетевые решения, средства связи': '51',
    'Тендеры': '53',
    'Товары народного потребления': '55',
    'Услуги для бизнеса': '57',
    'Флористика': '59',
    'Электротехническое оборудование, светотехника': '61',
    'Ювелирные изделия': '63',
    'Другое': '65',
    'Начало карьеры, мало опыта': '67',
}

# Искусство, культура, развлечения
xparh_kultura = {


'Балет, хореография, танцы':'1',
'Библиотеки':'3',
'Декоративно-прикладное искусство, народные промыслы':'5',
'Живопись, графика, скульптура':'7',
'Искусствоведение':'9',
'Кино, мультипликация':'11',
'Литература':'13',
'Модельный бизнес':'15',
'Музей, выставочный зал':'17',
'Музыка, пение':'19',
'Развлечения, игры, отдых, анимация':'21',
'Реставрация, антиквариат':'23',
'Театр':'25',
'Шоу-бизнес':'27',
'Другое':'29',
'Начало карьеры, мало опыта':'31',

}

# Кадры, управление персоналом
xparh_hr_kadri = {


'Кадровый учет, делопроизводство, кадровое администрирование':'1',
'Оплата труда, компенсации и льготы':'3',
'Охрана труда':'5',
'Психология труда и социальная психология':'7',
'Развитие персонала, обучение, тренинги':'9',
'Рекрутмент, подбор персонала':'11',
'Управление персоналом':'13',
'Другое':'15',
'Начало карьеры, мало опыта':'17',
}

# Консалтинг, стратегическое развитие
xparh_konsalting = {

    'IT-консалтинг': '1',
    'Кадровый консалтинг': '3',
    'Корпоративные финансы': '5',
    'Маркетинговый и PR консалтинг': '7',
    'Реинжиниринг бизнес-процессов': '9',
    'Стратегический консалтинг': '11',
    'Управление проектами': '13',
    'Управленческий консалтинг': '15',
    'Экологический консалтинг': '17',
    'Другое': '19',
    'Начало карьеры, мало опыта': '21',

}

# Маркетинг, реклама, PR
xparh_marketing = {

    'BTL': '1',
    'Event, организация мероприятий': '3',
    'GR, IR': '5',
    'PR': '7',
    'SMM': '9',
    'Аналитика, исследование рынков': '11',
    'Бренд-менеджмент, продакт-менеджмент': '13',
    'Интернет-маркетинг': '15',
    'Копирайтинг, редактирование': '17',
    'Маркетинг': '19',
    'Медиаисследования, рейтинги': '21',
    'Медиапланирование': '23',
    'Мерчандайзинг': '25',
    'Наружная реклама': '27',
    'Печатная реклама': '29',
    'Политический PR': '31',
    'Продвижение, специальные мероприятия': '33',
    'Производство рекламы': '35',
    'Радиореклама': '37',
    'Рекламно-сувенирная продукция': '39',
    'Рекламное агентство': '41',
    'Социологические исследования': '43',
    'Телевизионная реклама': '45',
    'Телемаркетинг': '47',
    'Торговый маркетинг': '49',
    'Другое': '51',
    'Начало карьеры, мало опыта': '53',

}

# Медицина, фармацевтика, ветеринария
xparh_medicina = {

    'Больницы, поликлиники, диагностические центры, лаборатории': '1',
    'Ветеринария': '3',
    'Клинические испытания': '5',
    'Лечебное дело (врачи-специалисты)': '7',
    'Лицензирование': '9',
    'Медицинское оборудование': '11',
    'Санитарно-эпидемиологический надзор': '13',
    'Сертификация': '15',
    'Средний и младший медицинский персонал': '17',
    'Товары медицинского назначения': '19',
    'Уход за больными': '21',
    'Фармацевтика': '23',
    'Другое': '25',
    'Начало карьеры, мало опыта': '27',

}

# Наука, образование, повышение квалификации
xparh_obrazovanie = {

    'Академия наук': '1',
    'Бизнес-образование': '3',
    'Внешкольное образование': '5',
    'Высшее образование': '7',
    'Дошкольное воспитание и образование': '9',
    'Инновационные технологии': '11',
    'Курсы, тренинги, семинары, повышение квалификации': '13',
    'НИИ, КБ': '15',
    'Репетиторство': '17',
    'Среднее образование': '19',
    'Среднее специальное образование': '21',
    'Другое': '23',
    'Начало карьеры, мало опыта': '25',

}


#Некоммерческие организации, волонтерство
xpath_nekom_org ={
'Благотворительность':'1',
'Общественные организации, ассоциации, фонды':'3',
'Религиозные организации':'5',
'Другое':'7',
'Начало карьеры, мало опыта':'9',

}
#Продажи
xpath_prodazi ={
'Call Center':'1',
'Автомобили, запчасти':'3',
'Алкоголь, напитки':'5',
'Бытовая техника, электроника, фото, видео':'7',
'Вендинг':'9',
'Госзакупки':'11',
'ГСМ, нефть':'13',
'Зоотовары, ветеринарные препараты':'15',
'Канцелярские товары':'17',
'Книги, печатные издания':'19',
'Компьютеры, оргтехника, ПО':'21',
'Косметика, бытовая химия':'23',
'Логистика':'25',
'Мебель':'27',
'Медицина, фармацевтика':'29',
'Металлопрокат':'31',
'Оборудование':'33',
'Оптовая торговля':'35',
'Продукты питания':'37',
'Продукция химического производства':'39',
'Рекламно-сувенирная продукция':'41',
'Сантехника':'43',
'Сельское хозяйство':'45',
'Системы безопасности':'47',
'Спортивные товары и фитнес-услуги':'49',
'Строительно-отделочные материалы':'51',
'Сырье':'53',
'Табачная продукция':'55',
'Текстиль, одежда, обувь':'57',
'Телекоммуникации, сетевые решения, средства связи':'59',
'Тендеры':'61',
'Товары народного потребления':'63',
'Услуги для бизнеса':'65',
'Флористика':'67',
'Электротехническое оборудование, светотехника':'69',
'Ювелирные изделия':'71',
'Другое':'73',
'Начало карьеры, мало опыта':'75',

}
#Промышленность, производство
xpath_promishlen ={

'Авиационная промышленность':'1',
'Автомобильная промышленность':'3',
'Атомная энергетика':'5',
'Железнодорожное машиностроение':'7',
'Контроль качества, сертификация, экспертиза':'9',
'Легкая промышленность':'11',
'Лесная промышленность':'13',
'Машиностроение, станкостроение':'15',
'Мебельное производство':'17',
'Металлургия':'19',
'Нефтегазовое машиностроение':'21',
'Нефтепереработка':'23',
'Пищевая промышленность':'25',
'Полиграфическое производство':'27',
'Приборостроение':'29',
'Производство алкогольных и безалкогольных напитков':'31',
'Производство металлических изделий и заготовок':'33',
'Производство стройматериалов':'35',
'Производство товаров народного потребления':'37',
'Промышленное оборудование':'39',
'Радиотехническая и электронная промышленность':'41',
'Ракетно-космическая отрасль':'43',
'Робототехника':'45',
'Сельскохозяйственная техника':'47',
'Стекольная и фарфоро-фаянсовая промышленность':'49',
'Судостроение':'51',
'Химическая промышленность':'53',
'Химическое машиностроение':'55',
'Экология':'57',
'Энергетика':'59',
'Энергомашиностроение':'61',
'Ювелирная промышленность':'63',
'Другое':'65',
'Начало карьеры, мало опыта':'67',
}
#Рабочий персонал
xpath_rab_person ={

'Бригада':'1',
'Квалифицированный рабочий':'3',
'Рабочий персонал, разное':'5',

}
#СМИ, издательства
xpath_smi_izdatel ={

'Журналистика':'1',
'Издательская деятельность':'3',
'Интернет издания':'5',
'Литературные переводы':'7',
'Печатные издания':'9',
'Радио':'11',
'Редактура, корректура':'13',
'Телевидение':'15',
'Другое':'17',
'Начало карьеры, мало опыта':'19',
}
#Сельское хозяйство
xpath_selskoe ={

'Животноводство':'1',
'Птицеводство':'3',
'Растениеводство':'5',
'Рыбоводство':'7',
'Другое':'9',
'Начало карьеры, мало опыта':'11',
}
#Спорт, фитнес, салоны красоты, SPA
xpath_sport ={
'Косметология':'1',
'Маникюр, педикюр':'3',
'Массаж':'5',
'Парикмахерское дело':'7',
'Салоны красоты / Парикмахерские / SPA':'9',
'Стилистика, визаж':'11',
'Тренерская работа, инструктаж':'13',
'Другое':'15',
'Начало карьеры, мало опыта':'17',
}
#Страхование
xpath_strahovanie ={
'Автострахование':'1',
'Актуарная деятельность':'3',
'Андеррайтинг':'5',
'Медицинское страхование':'7',
'Перестрахование':'9',
'Страхование бизнеса':'11',
'Страхование грузов':'13',
'Страхование жизни':'15',
'Страхование имущества':'17',
'Страхование коммерческих и финансовых рисков':'19',
'Страхование наружной рекламы':'21',
'Страхование недвижимости, ипотечное страхование':'23',
'Страхование ответственности':'25',
'Страхование строительно-монтажных работ':'27',
'Страхование туристов':'29',
'Страхование физических лиц':'31',
'Страхование юридических лиц':'33',
'Страховой анализ, оценка':'35',
'Урегулирование убытков':'37',
'Другое':'39',
'Начало карьеры, мало опыта':'41',
}
#Строительство, проектирование, недвижимость
xpath_stroitelstvo ={
'Агентства недвижимости, оценка недвижимости':'1',
'Архитектура':'3',
'Благоустройство территорий':'5',
'Водоснабжение, канализация':'7',
'Газоснабжение':'9',
'Геодезия, картография, землеустроительство':'11',
'ЖКХ':'13',
'Конструирование':'15',
'Отопление, вентиляция и кондиционирование':'17',
'Подводно-технические работы':'19',
'Проектирование':'21',
'Разработка генерального плана':'23',
'Слаботочные системы и сети':'25',
'Сметное дело':'27',
'Строительно-монтажные и отделочные работы':'29',
'Технический надзор, строительная экспертиза':'31',
'Транспортные системы (мосты, дороги, тоннели)':'33',
'Трубопроводы':'35',
'Управление недвижимостью':'37',
'Управление проектами (ГИП)':'39',
'Управление строительством, девелопмент':'41',
'Экологическое сопровождение проектов':'43',
'Эксплуатация зданий':'45',
'Электроснабжение':'47',
'Другое':'49',
'Начало карьеры, мало опыта':'51',
}
#Сырье
xpath_sirje ={
'Газ':'1',
'Добыча':'3',
'Металлы':'5',
'Нефть':'7',
'Оборудование':'9',
'Переработка':'11',
'Разведка и разработка месторождений':'13',
'Технологии':'15',
'Уголь':'17',
'Другое':'19',
'Начало карьеры, мало опыта':'21',
}
#Топ-персонал
xpath_top_personal ={
'IT, Интернет, связь, телеком':'1',
'Административная работа, секретариат, АХО':'3',
'Банки, инвестиции, лизинг':'5',
'Безопасность, службы охраны':'7',
'Бухгалтерия, финансы, аудит':'9',
'Государственная служба, некоммерческие организации':'11',
'Дизайн':'13',
'Закупки, снабжение':'15',
'Искусство, культура, развлечения':'17',
'Кадры, управление персоналом':'19',
'Консалтинг, стратегическое развитие':'21',
'Маркетинг, реклама, PR':'23',
'Медицина, фармацевтика, ветеринария':'25',
'Наука, образование, повышение квалификации':'27',
'Продажи':'29',
'Промышленность, производство, сельское хозяйство':'31',
'СМИ, издательство, полиграфия':'33',
'Спорт, фитнес, салоны красоты, SPA':'35',
'Страхование':'37',
'Строительство, проектирование, недвижимость':'39',
'Сырье':'41',
'Транспорт, логистика, ВЭД':'43',
'Туризм, гостиницы, общественное питание':'45',
'Услуги, ремонт, сервисное обслуживание':'47',
'Юриспруденция':'49',
}
#Транспорт, логистика, ВЭД
xpath_transport ={

'Авиаперевозки':'1',
'Автоперевозки':'3',
'ВЭД':'5',
'Железнодорожные перевозки':'7',
'Контейнерные перевозки':'9',
'Логистика':'11',
'Метрополитен':'13',
'Морские, речные перевозки':'15',
'Складское хозяйство':'17',
'Таможня':'19',
'Трубопроводы':'21',
'Другое':'23',
'Начало карьеры, мало опыта':'25',
}
#Туризм, гостиницы, общественное питание
xpath_turizm ={

'HORECA':'1',
'Гостиницы, отели, кемпинги, мотели':'3',
'Кейтеринг':'5',
'Рестораны, кафе, столовые, фастфуд':'7',
'Туристические услуги и продукты':'9',
'Другое':'11',
'Начало карьеры, мало опыта':'13',
}
#Услуги, ремонт, сервисное обслуживание
xpath_remont ={

'Автосервис':'1',
'Ателье':'3',
'Бани':'5',
'Бытовая техника':'7',
'Клининг':'9',
'Компьютеры и оргтехника':'11',
'Медицинская техника':'13',
'Металлоремонт':'15',
'Оборудование для индустрии развлечения и спортивного инвентаря':'17',
'Обувь':'19',
'Промышленное оборудование':'21',
'Ритуальные услуги':'23',
'Телекоммуникационное оборудование и сети':'25',
'Торговое, складское, холодильное оборудование':'27',
'Фото, аудио, видео услуги':'29',
'Химчистки, прачечные':'31',
'Другое':'33',
'Начало карьеры, мало опыта':'35',
}
#Юриспруденция
xpath_urizm = {
'Compliance (нормативное право, соблюдение закона )':'1',
'Авторское право, патентное право':'3',
'Адвокатские услуги':'5',
'Антимонопольное право':'7',
'Арбитраж':'9',
'Банковское право':'11',
'Военное право':'13',
'Гражданское право':'15',
'Договорное право':'17',
'Законотворчество':'19',
'Земельное право':'21',
'Имущественное право':'23',
'Интеллектуальная собственность':'25',
'Корпоративное право':'27',
'Лицензирование':'29',
'Медицинское право':'31',
'Международное право':'33',
'Морское право':'35',
'Налоговое право':'37',
'Недропользование':'39',
'Нотариат':'41',
'Регистрация, перерегистрация, ликвидация предприятий':'43',
'Семейное право':'45',
'Слияния и поглощения':'47',
'Страховое право':'49',
'Таможенное право':'51',
'Трудовое право':'53',
'Уголовное право':'55',
'Урегулирование убытков':'57',
'Финансовое право':'59',
'Ценные бумаги':'61',
'Экологическое право':'63',
'Юридические консультации и услуги':'65',
'Другое':'67',
'Начало карьеры, мало опыта':'69',

}

vibor_podrubrik = {
    'IT, Интернет, связь, телеком' : xpath_it_ithernet,
    'Административная работа, секретариат, АХО' : xpath_it_adm_personal,
    'Банки, инвестиции, лизинг' : xpath_banki_invest,
    'Транспорт, логистика, ВЭД' : xpath_logistika,
    'Безопасность, службы охраны' : xpath_ohrana,
    'Бухгалтерия, финансы, аудит' : xparh_buhgalteria,
    'Дизайн' : xparh_design,
    'Домашний персонал' : xparh_home_personal,
    'Закупки, снабжение' : xparh_zakupki,
    'Исскуство, культура, развлечения' : xparh_kultura,
    'Кадры, управление персоналом' : xparh_hr_kadri,
    'Консалтинг, стратегическое развитие' : xparh_konsalting,
    'Маркетинг, реклама, PR' : xparh_marketing,
    'Медицина,  фармацевтика, ветеринария' : xparh_medicina,
    'Наука, образование, повышение квалификации' : xparh_obrazovanie,
    'Некоммерческие организации, волонтерство':xpath_nekom_org,
    'Продажи':xpath_prodazi,
    'Промышленность, производство':xpath_promishlen,
    'Рабочий персонал':xpath_rab_person,
    'СМИ, издательства':xpath_smi_izdatel,
    'Сельское хозяйство':xpath_selskoe,
    'Спорт, фитнес, салоны красоты, SPA':xpath_sport,
    'Страхование':xpath_strahovanie,
    'Строительство, проектирование, недвижимость':xpath_stroitelstvo,
    'Сырье':xpath_sirje,
    'Топ-персонал':xpath_top_personal,
    'Транспорт, логистика, ВЭД':xpath_transport,
    'Туризм, гостиницы, общественное питание':xpath_turizm,
    'Услуги, ремонт, сервисное обслуживание':xpath_remont,
    'Юриспруденция':xpath_urizm

    }
################### считываем данные с excel ################################

wb = xlrd.open_workbook('./DATA.xlsx')
sheet_xlr = wb.sheet_by_name('Вакансии')
a = 2


rows_sheet = sheet_xlr.nrows
# print(rows_sheet)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)

sheet = wb[lst[0]]
sheet.title
error_string = 1056

# random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.

error_string = 1060
i = random.randint(3, rows_sheet)
################### заводим содержимое в переменные ################################
# print(a)
vacancy = (sheet.cell(row=i, column=1).value)
zp_ot = (sheet.cell(row=i, column=2).value)
zp_do = (sheet.cell(row=i, column=3).value)
opyt = (sheet.cell(row=i, column=4).value)
opisanie = (sheet.cell(row=i, column=5).value)
grafik = (sheet.cell(row=i, column=6).value)
company = (sheet.cell(row=i, column=7).value)
opis_company = (sheet.cell(row=i, column=8).value)
company1 = (sheet.cell(row=i, column=9).value)
opis_company1 = (sheet.cell(row=i, column=10).value)
company2 = (sheet.cell(row=i, column=11).value)
opis_company2 = (sheet.cell(row=i, column=12).value)
company3 = (sheet.cell(row=i, column=13).value)
opis_company3 = (sheet.cell(row=i, column=14).value)
company4 = (sheet.cell(row=i, column=15).value)
opis_company4 = (sheet.cell(row=i, column=16).value)
sfera_deyatel = (sheet.cell(row=i, column=17).value)
opis_deyatel = (sheet.cell(row=i, column=18).value)
name_company = (sheet.cell(row=i, column=19).value)
opis_deyatel_rename_company = (sheet.cell(row=i, column=20).value)

# print(vacancy)
error_string = 1099
# СЧИТЫВАЕМ ГОРОДА иЗ ЭКСЕЛЯ ДИНАМИЧЕСКИ

sheet = wb[lst[1]]
sheet.title

barnaul_adr = []
volgograd_adr = []
voronej_adr = []
ekaterenburg_adr = []
izevsk_adr = []
irkuts_adr = []
kazan_adr = []
kaliningrad_adr = []
kemerovo_adr = []
krasnodar_adr = []
krasnoyars_adr = []
moskow_adr = []
naber_cheln_adr = []
nizniy_novgorod_adr = []
novosib_adr = []
omsk_adr = []
orenburg_adr = []
perm_adr = []
rostov_na_dony_adr = []
ryazan_adr = []
samara_adr = []
sankt_peter_adr = []
saratov_adr = []
sochi_and_adler = []
toliatty_adr = []
tomsk_adr = []
tyla_adr = []
tumen_adr = []
ufa_adr = []
chelabinsk_adr = []
yaroslavl_adr = []
musor_exela = []
#goroda_arr = [musor_exela, barnaul_adr, volgograd_adr, voronej_adr, ekaterenburg_adr, izevsk_adr, irkuts_adr, kazan_adr, kaliningrad_adr, kemerovo_adr, krasnodar_adr, krasnoyars_adr, moskow_adr, naber_cheln_adr, nizniy_novgorod_adr, novosib_adr, omsk_adr, orenburg_adr, perm_adr, rostov_na_dony_adr, ryazan_adr, samara_adr, sankt_peter_adr, saratov_adr, sochi_and_adler, toliatty_adr, tomsk_adr, tyla_adr, tumen_adr, ufa_adr, chelabinsk_adr, yaroslavl_adr]
goroda_arr = []

kolichestvo_gorodov = 0

for cell in sheet['A']:
    kolichestvo_gorodov += 1
    print(cell.value)

for number_gorod in range(kolichestvo_gorodov):
    goroda_arr.append([])
    for cell in list(sheet.rows)[number_gorod]:
         if cell.value != None:
            goroda_arr[number_gorod].append(cell.value)

kolichestvo_adresov = []
high_number_gorod = 0
number_goroda = 0
while number_goroda < len(goroda_arr):
    max_number = len(goroda_arr[number_goroda])
    kolichestvo_adresov.append(len(goroda_arr[number_goroda]))
    if high_number_gorod< max_number:
        high_number_gorod = max_number
        number_goroda = number_goroda + 1
    else:
        number_goroda = number_goroda + 1

###############    autorith  ###############

n = 0
while n < 1500:
    sheet = wb[lst[2]]
    sheet.title
    login = (sheet.cell(row=1, column=2).value)
    password = (sheet.cell(row=2, column=2).value)

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    browser = webdriver.Chrome(chrome_options=options)
    driver = browser

    browser.get('https://nn.superjob.ru/hr/vacancy/create/')



    ############### Login ########################

    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[3]/div/div/div/div/div[1]/label/div/div/input').send_keys(login)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[3]/div/div/div/div/div[2]/label/div/div[1]/input').send_keys(password)
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[4]/div/div[1]/button').click()

    ############### Должность ####################
    time.sleep(4)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[1]/div/div[2]/div/div[1]/label/div/div/input').send_keys(vacancy)

    ############## Специализация #################
    companys = [company, company1, company2, company3, company4]
    opis_companys = [opis_company, opis_company1, opis_company2, opis_company3, opis_company4]
    iter_companys = 0 #Для теста пока без цикла, потом добавить как в работару
    print(companys)
    print(opis_companys)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[2]/div/div/div/button').click()
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[3]/div/button/span/span/span').click()
    except:
        print('чисто')
    time.sleep(1)
    while iter_companys < 5:
        if iter_companys != 0:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[2]/div/div/div/button').click()
        #Тыкаем на специализацию!
        specializacii = browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div/ul')
        specializacii_list = specializacii.find_elements_by_tag_name('span')
        n_spec = xpath_specializacii.get(companys[iter_companys])
        specializacii_list[int(n_spec)].click()
        time.sleep(1)
        #тыкаем на подкатегорию
        xpartt = vibor_podrubrik.get(companys[iter_companys])
        xpartt2 = xpartt.get(opis_companys[iter_companys])
        oblast_specializacii = browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[2]/div/div[3]/ul')
        speciall = oblast_specializacii.find_elements_by_tag_name('span')
        speciall[int(xpartt2)].click()
        time.sleep(1)
        browser.find_element_by_xpath('/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[3]/div/div/div[1]/button/span/span/span').click()
        time.sleep(1)

        iter_companys+=1

    ############# Занятость #################### НУЖНО СДЕЛАТЬ

    browser.find_element_by_xpath('//*[@id="detailInfo.workType.id-input"]').click()
    browser.find_element_by_xpath('//*[@id="detailInfo.workType.id-input"]').send_keys(grafik + Keys.ENTER)
    ############# Уровень дохода ###############

    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[7]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[1]/label/div/div[2]/input').send_keys(zp_ot)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[7]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[2]/label/div/div[2]/input').send_keys(zp_do)

    ############ Опыт работы ###################
    try:
        if opyt == 'не имеет значения':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[1]').click()
        elif opyt == '1 год':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[3]').click()
        elif opyt == '3 года':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[4]').click()
        elif opyt == '6 лет':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[5]').click()
        elif opyt == 'без опыта':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[2]').click()
    except:
        print('Выбраный опыт работы уже стоит')

    ############ Описание вакансии #############

    #Вытаскиваем Условия
    frag_opisanie = opisanie.partition('Условия:')
    frag_opisanie = frag_opisanie[2].partition('Обязанности:')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    yslovia = frag_opisanie[0]

    #Вытаскиваем Обязанности
    frag_opisanie = opisanie.partition('Обязанности:')
    frag_opisanie = frag_opisanie[2].partition('Требования:')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    obiaznosti = frag_opisanie[0]

    #Вытаскиваем Требования
    frag_opisanie = opisanie.partition('Требования:')
    frag_opisanie = frag_opisanie[2].partition('\n \n')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    trebovania = frag_opisanie[0]

    #Вытаскиваем шапку
    frag_opisanie = opisanie.partition('Условия:')
    shapka = frag_opisanie[0]
    print(frag_opisanie)

    #Вытаскиваем подвал
    frag_opisanie = opisanie.partition('\n \n')
    podval = frag_opisanie[2]

    #Заполняем описание
    linii = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[10]/div/div/div[2]/div/div[1]/div/div/div/div[2]')
    linii.find_element_by_tag_name('b').send_keys(Keys.ARROW_UP + Keys.ENTER)
    lst_linii = linii.find_elements_by_tag_name('li')
    linii.find_element_by_tag_name('p').send_keys(Keys.LEFT_CONTROL + 'b' + Keys.LEFT_CONTROL + (shapka + Keys.BACKSPACE))
    lst_linii[0].send_keys(obiaznosti + Keys.BACKSPACE + Keys.DELETE)
    lst_linii[2].send_keys(trebovania + Keys.DELETE)
    lst_linii[4].send_keys(yslovia + Keys.BACKSPACE + Keys.DELETE)
    linii.send_keys('\n\n' + Keys.LEFT_CONTROL + 'b' + Keys.LEFT_CONTROL + podval)



    #Заполняем города
    #ля теста iter_gorod
    iter_gorod = 1
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[1]/div/div/div[2]/ul/li[1]/span/span[2]/button').click()
    except:
        print('город автоматически не проставился')

    adr_bar = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[1]/div/div/div[3]/div/div[1]/label/div/div/input')
    random_adres = random.randint(1, kolichestvo_adresov[iter_gorod]-1)
    print('random adres = ', random_adres)
    adr_bar.send_keys(goroda_arr[iter_gorod][0])

    adr_bar = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[2]/div/div/div/div[2]/div/div[1]/div/div[1]/label/div/div/input')
    adr_bar.send_keys(goroda_arr[iter_gorod][random_adres])
    time.sleep(2)
    adr_bar.send_keys(Keys.ARROW_DOWN + Keys.ENTER)

    #########Добавляем метро
    try:
        browser.find_element_by_xpath(
            '/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[5]/div/div/div/button/span/span/span').click()
        time.sleep(2)
        list_metro = browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]')
        spisoc_metro = list_metro.find_elements_by_class_name('_2s93E')

        metro = len(spisoc_metro) - 1
        array_metro = list(range(0, metro))
        random.shuffle(array_metro)
        iter_metro = 0
        while iter_metro < 10:
            iter_iter_metro = array_metro[iter_metro]
            spisoc_metro[iter_iter_metro].click()
            time.sleep(1)
            iter_metro += 1
        browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[3]/div/button').click()
    except:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[3]/div/button').click()
    ############## Информация о компании ###################################

    if sfera_deyatel != None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/label/div/div[1]/span').click()
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[2]/button').click()
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[1]/div/label/div/div/input').send_keys(sfera_deyatel)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[2]/div/label/div/div/textarea').send_keys(opis_deyatel)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[3]/div/div[1]/button').click()
        except:
            print("Скрыть информацию не получилось")
    elif name_company != None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[2]/button').click()
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[1]/div/label/div/div/input').send_keys(name_company)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[2]/div/label/div/div/textarea').send_keys(opis_deyatel_rename_company)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[3]/div/div[1]/button').click()
        except:
            print("Отредактировать информацию не получилось")
    else:
        print("Оставляем название компании ткаим какое оно есть")


    ############## Размещаем вакансию и закрываем браузер ##################
    time.sleep(1)
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[6]/div/div[2]/div/div[1]/div/div[1]/button/span/span/span').click()
    except:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[6]/div/div[2]/div/div[1]/div/div[1]/button')
    time.sleep(1)
    browser.quit()
    n+=1


#browser.quit()