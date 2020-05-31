import time
import selenium
from selenium import webdriver
from openpyxl import load_workbook
import random
import xlrd
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

################### Xpath Для специализации ################################

xpath_specializacii = {

'Административная работа, секретариат, АХО':'0',
'Бухгалтерия, финансы, аудит':'1',
'IT, Интернет, связь, телеком':'2',
'Дизайн':'3',
'Кадры, управление персоналом':'4',
'Транспорт, логистика, ВЭД':'5',
'Юриспруденция':'6',
'Медицина, фармацевтика, ветеринария':'7',
'Некоммерческие организации, волонтерство':'8',
'Безопасность, службы охраны':'9',
'Туризм, гостиницы, общественное питание':'10',
'Искусство, культура, развлечения':'11',
'СМИ, издательства':'12',
'Маркетинг, реклама, PR':'13',
'Спорт, фитнес, салоны красоты, SPA':'14',
'Наука, образование, повышение квалификации':'15',
'Страхование':'16',
'Строительство, проектирование, недвижимость':'17',
'Промышленность, производство':'18',
'Услуги, ремонт, сервисное обслуживание':'19',
'Банки, инвестиции, лизинг':'20',
'Сырье':'21',
'Консалтинг, стратегическое развитие':'22',
'Продажи':'23',
'Домашний персонал':'24',
'Топ-персонал':'25',
'Рабочий персонал':'26',
'Закупки, снабжение':'27',
'Сельское хозяйство':'28',
}

################### Xpath Для подрубрик ################################

# IT / Интернет / Телеком
xpath_it_ithernet = {

'IT, Интернет, связь, телеком':'0',
'CRM-системы':'2',
'Web, UI, UX дизайн':'4',
'Web-верстка':'8',
'Администрирование баз данных':'10',
'Аналитика':'14',
'Внедрение и сопровождение ПО':'16',
'Защита информации':'20',
'Игровое ПО / Геймдевелопмент':'22',
'Инжиниринг':'26',
'Интернет, создание и поддержка сайтов':'28',
'Киберспорт':'32',
'Компьютерная анимация и мультимедиа':'34',
'Контент':'38',
'Мобильная разработка':'40',
'Оптимизация, SEO':'44',
'Передача данных и доступ в интернет':'46',
'Разработка и сопровождение банковского ПО':'50',
'Разработка, программирование':'52',
'Сетевые технологии':'56',
'Системная интеграция':'58',
'Системное администрирование':'62',
'Системы автоматизированного проектирования (САПР)':'64',
'Системы управления предприятием (ERP)':'68',
'Сотовые, беспроводные технологии':'70',
'Телекоммуникации и связь':'74',
'Тестирование, QA':'76',
'Техническая документация':'80',
'Техническая поддержка':'82',
'Управление продуктом':'86',
'Управление проектами':'88',
'Электронная коммерция':'92',
'Электронный документооборот':'94',
'Юзабилити':'98',
'Другое':'100',
'Начало карьеры, мало опыта':'104',

}

# Административная работа, секретариат, АХО
xpath_it_adm_personal = {

'Архивное дело':'2',
'АХО':'4',
'Делопроизводство, ввод данных, систематизация':'8',
'Диспетчерская служба':'10',
'Курьерская служба':'14',
'Переводы':'16',
'Секретариат, ресепшн, офис-менеджмент':'20',
'Другое':'22',
'Начало карьеры, мало опыта':'26',

}

# Транспорт, логистика, ВЭД
xpath_logistika = {

'Авиаперевозки':'1',
'Автоперевозки':'3',
'ВЭД':'5',
'Железнодорожные перевозки':'7',
'Контейнерные перевозки':'9',
'Логистика':'11',
'Метрополитен':'13',
'Морские, речные перевозки':'15',
'Складское хозяйство':'17',
'Таможня':'19',
'Трубопроводы':'21',
'Другое':'23',
'Начало карьеры, мало опыта':'25',
}

# Банки, инвестиции, лизинг
xpath_banki_invest = {

'Банковская бухгалтерия':'2',
'Бэк-Офис':'4',
'Бюджетирование и планирование':'8',
'Валютные операции':'10',
'Вклады':'14',
'Депозитарий':'16',
'Документарные операции':'20',
'Залоги и проблемная задолженность':'22',
'Ипотека':'26',
'Кредитование физических лиц':'28',
'Кредитование юридических лиц':'32',
'Лизинг':'34',
'Методология, разработка и продажа корпоративных продуктов':'38',
'Методология, разработка и продажа продуктов Private Banking':'40',
'Методология, разработка и продажа розничных продуктов':'44',
'Налоговый учёт':'46',
'Обслуживание банкоматов':'50',
'Операции с драгметаллами':'52',
'Пластиковые карты (эквайринг)':'56',
'Продажа банковских продуктов':'58',
'Разработка банковских продуктов':'62',
'Расчёты и обработка платежей, касса':'64',
'Риски':'68',
'Торговое финансирование':'70',
'Управление активами':'74',
'Управление ликвидностью и балансовыми рисками':'76',
'Управленческая отчетность':'80',
'Финансовая отчётность':'82',
'Финансовый анализ и контроль':'86',
'Ценные бумаги':'88',
'Другое':'92',
'Начало карьеры, мало опыта':'94',
}

# Безопасность, службы охраны
xpath_ohrana= {

'Видеонаблюдение':'2',
'Имущественная безопасность':'4',
'Инкассация':'8',
'Информационная безопасность':'10',
'Кинология':'14',
'Личная безопасность':'16',
'Охранно-, детективная деятельность':'20',
'Охранные службы предприятий':'22',
'Пожарная безопасность':'26',
'Служба спасения':'28',
'ЧОП':'32',
'Экономическая безопасность':'34',
'Другое':'38',
'Начало карьеры, мало опыта':'40',
    }

# Бухгалтерия, финансы, аудит
xparh_buhgalteria = {

'GAAP':'2',
'Аудит, ревизия, инспектирование':'4',
'Бухгалтерский учет':'8',
'Бюджетирование и планирование':'10',
'Казначейство':'14',
'Кассовые операции':'16',
'Кредитный контроль':'20',
'МСФО, IFRS, РСБУ':'22',
'Налоги':'26',
'Основные средства':'28',
'Первичная документация':'32',
'Планово-экономическое управление':'34',
'Расчет себестоимости':'38',
'Учет заработной платы':'40',
'Учет счетов и платежей':'44',
'Учет труда и социальных выплат сотрудникам':'46',
'Учет фондов':'50',
'Финансовый анализ и контроль':'52',
'Финансовый менеджмент':'56',
'Экономический анализ':'58',
'Другое':'62',
'Начало карьеры, мало опыта':'64',

}

# Дизайн
xparh_design = {

'Архитектура':'2',
'Аудио, видео':'4',
'Верстка, дизайн':'8',
'Графическое искусство, иллюстрации':'10',
'Дизайн интерьера':'14',
'Компьютерная анимация и мультимедиа':'16',
'Ландшафтный дизайн':'20',
'Мода и Аксессуары':'22',
'Промышленный дизайн':'26',
'Рекламный дизайн':'28',
'Фотография':'32',
'Другое':'34',
'Начало карьеры, мало опыта':'38',

}

# Домашний персонал
xparh_home_personal = {

'Ведение хозяйства':'2',
'Домашний уход за больными, престарелыми, инвалидами':'4',
'Домашний уход за детьми':'8',
'Уборка и помощь по дому':'10',
'Другое':'14',
'Начало карьеры, мало опыта':'16',

}

# Закупки, снабжение
xparh_zakupki = {

'Автомобили, запчасти':'2',
'Алкоголь, напитки':'4',
'Бытовая техника, электроника, фото, видео':'8',
'Вендинг':'10',
'Госзакупки':'14',
'ГСМ, нефть':'16',
'Зоотовары, ветеринарные препараты':'20',
'Канцелярские товары':'22',
'Книги, печатные издания':'26',
'Компьютеры, оргтехника, ПО':'28',
'Косметика, бытовая химия':'32',
'Мебель':'34',
'Медицина, фармацевтика':'38',
'Металлопрокат':'40',
'Оборудование':'44',
'Продукты питания':'46',
'Продукция химического производства':'50',
'Сантехника':'52',
'Сельское хозяйство':'56',
'Системы безопасности':'58',
'Спортивные товары и фитнес-услуги':'62',
'Строительно-отделочные материалы':'64',
'Сырье':'68',
'Табачная продукция':'70',
'Текстиль, одежда, обувь':'74',
'Телекоммуникации, сетевые решения, средства связи':'76',
'Тендеры':'80',
'Товары народного потребления':'82',
'Услуги для бизнеса':'86',
'Флористика':'88',
'Электротехническое оборудование, светотехника':'92',
'Ювелирные изделия':'94',
'Другое':'98',
'Начало карьеры, мало опыта':'100',
}

# Искусство, культура, развлечения
xparh_kultura = {


'Балет, хореография, танцы':'2',
'Библиотеки':'4',
'Декоративно-прикладное искусство, народные промыслы':'8',
'Живопись, графика, скульптура':'10',
'Искусствоведение':'14',
'Кино, мультипликация':'16',
'Литература':'20',
'Модельный бизнес':'22',
'Музей, выставочный зал':'26',
'Музыка, пение':'28',
'Развлечения, игры, отдых, анимация':'32',
'Реставрация, антиквариат':'34',
'Театр':'38',
'Шоу-бизнес':'40',
'Другое':'44',
'Начало карьеры, мало опыта':'46',

}

# Кадры, управление персоналом
xparh_hr_kadri = {

'Кадровый учет, делопроизводство, кадровое администрирование':'2',
'Оплата труда, компенсации и льготы':'4',
'Охрана труда':'8',
'Психология труда и социальная психология':'10',
'Развитие персонала, обучение, тренинги':'14',
'Рекрутмент, подбор персонала':'16',
'Управление персоналом':'20',
'Другое':'22',
'Начало карьеры, мало опыта':'26',

}

# Консалтинг, стратегическое развитие
xparh_konsalting = {

'IT-консалтинг':'2',
'Кадровый консалтинг':'4',
'Корпоративные финансы':'8',
'Маркетинговый и PR консалтинг':'10',
'Реинжиниринг бизнес-процессов':'14',
'Стратегический консалтинг':'16',
'Управление проектами':'20',
'Управленческий консалтинг':'22',
'Экологический консалтинг':'26',
'Другое':'28',
'Начало карьеры, мало опыта':'32',

}

# Маркетинг, реклама, PR
xparh_marketing = {

'BTL':'2',
'Event, организация мероприятий':'4',
'GR, IR':'8',
'PR':'10',
'SMM':'14',
'Аналитика, исследование рынков':'16',
'Бренд-менеджмент, продакт-менеджмент':'20',
'Интернет-маркетинг':'22',
'Копирайтинг, редактирование':'26',
'Маркетинг':'28',
'Медиаисследования, рейтинги':'32',
'Медиапланирование':'34',
'Мерчандайзинг':'38',
'Наружная реклама':'40',
'Печатная реклама':'44',
'Политический PR':'46',
'Продвижение, специальные мероприятия':'50',
'Производство рекламы':'52',
'Радиореклама':'56',
'Рекламно-сувенирная продукция':'58',
'Рекламное агентство':'62',
'Социологические исследования':'64',
'Телевизионная реклама':'68',
'Телемаркетинг':'70',
'Торговый маркетинг':'74',
'Другое':'76',
'Начало карьеры, мало опыта':'80',

}

# Медицина, фармацевтика, ветеринария
xparh_medicina = {

'Больницы, поликлиники, диагностические центры, лаборатории':'2',
'Ветеринария':'4',
'Клинические испытания':'8',
'Лечебное дело (врачи-специалисты)':'10',
'Лицензирование':'14',
'Медицинское оборудование':'16',
'Санитарно-эпидемиологический надзор':'20',
'Сертификация':'22',
'Средний и младший медицинский персонал':'26',
'Товары медицинского назначения':'28',
'Уход за больными':'32',
'Фармацевтика':'34',
'Другое':'38',
'Начало карьеры, мало опыта':'40',


}

# Наука, образование, повышение квалификации
xparh_obrazovanie = {

'Академия наук':'2',
'Бизнес-образование':'4',
'Внешкольное образование':'8',
'Высшее образование':'10',
'Дошкольное воспитание и образование':'14',
'Инновационные технологии':'16',
'Курсы, тренинги, семинары, повышение квалификации':'20',
'НИИ, КБ':'22',
'Репетиторство':'26',
'Среднее образование':'28',
'Среднее специальное образование':'32',
'Другое':'34',
'Начало карьеры, мало опыта':'38',

}


#Некоммерческие организации, волонтерство
xpath_nekom_org ={

'Благотворительность':'2',
'Общественные организации, ассоциации, фонды':'4',
'Религиозные организации':'8',
'Другое':'10',
'Начало карьеры, мало опыта':'14',

}
#Продажи
xpath_prodazi ={
'Call Center':'2',
'Автомобили, запчасти':'4',
'Алкоголь, напитки':'8',
'Бытовая техника, электроника, фото, видео':'10',
'Вендинг':'14',
'Госзакупки':'16',
'ГСМ, нефть':'20',
'Зоотовары, ветеринарные препараты':'22',
'Канцелярские товары':'26',
'Книги, печатные издания':'28',
'Компьютеры, оргтехника, ПО':'32',
'Косметика, бытовая химия':'34',
'Логистика':'38',
'Мебель':'40',
'Медицина, фармацевтика':'44',
'Металлопрокат':'46',
'Оборудование':'50',
'Оптовая торговля':'52',
'Продукты питания':'56',
'Продукция химического производства':'58',
'Рекламно-сувенирная продукция':'62',
'Сантехника':'64',
'Сельское хозяйство':'68',
'Системы безопасности':'70',
'Спортивные товары и фитнес-услуги':'74',
'Строительно-отделочные материалы':'76',
'Сырье':'80',
'Табачная продукция':'82',
'Текстиль, одежда, обувь':'86',
'Телекоммуникации, сетевые решения, средства связи':'88',
'Тендеры':'92',
'Товары народного потребления':'94',
'Услуги для бизнеса':'98',
'Флористика':'100',
'Электротехническое оборудование, светотехника':'104',
'Ювелирные изделия':'106',
'Другое':'110',
'Начало карьеры, мало опыта':'112',


}
#Промышленность, производство
xpath_promishlen ={

'Авиационная промышленность':'2',
'Автомобильная промышленность':'4',
'Атомная энергетика':'8',
'Железнодорожное машиностроение':'10',
'Контроль качества, сертификация, экспертиза':'14',
'Легкая промышленность':'16',
'Лесная промышленность':'20',
'Машиностроение, станкостроение':'22',
'Мебельное производство':'26',
'Металлургия':'28',
'Нефтегазовое машиностроение':'32',
'Нефтепереработка':'34',
'Пищевая промышленность':'38',
'Полиграфическое производство':'40',
'Приборостроение':'44',
'Производство алкогольных и безалкогольных напитков':'46',
'Производство металлических изделий и заготовок':'50',
'Производство стройматериалов':'52',
'Производство товаров народного потребления':'56',
'Промышленное оборудование':'58',
'Радиотехническая и электронная промышленность':'62',
'Ракетно-космическая отрасль':'64',
'Робототехника':'68',
'Сельскохозяйственная техника':'70',
'Стекольная и фарфоро-фаянсовая промышленность':'74',
'Судостроение':'76',
'Химическая промышленность':'80',
'Химическое машиностроение':'82',
'Экология':'86',
'Энергетика':'88',
'Энергомашиностроение':'92',
'Ювелирная промышленность':'94',
'Другое':'98',
'Начало карьеры, мало опыта':'100',
}
#Рабочий персонал
xpath_rab_person ={

'Бригада':'2',
'Квалифицированный рабочий':'4',
'Рабочий персонал, разное':'8',

}
#СМИ, издательства
xpath_smi_izdatel ={

'Журналистика':'2',
'Издательская деятельность':'4',
'Интернет издания':'8',
'Литературные переводы':'10',
'Печатные издания':'14',
'Радио':'16',
'Редактура, корректура':'20',
'Телевидение':'22',
'Другое':'26',
'Начало карьеры, мало опыта':'28',
}
#Сельское хозяйство
xpath_selskoe ={

'Животноводство':'2',
'Птицеводство':'4',
'Растениеводство':'8',
'Рыбоводство':'10',
'Другое':'14',
'Начало карьеры, мало опыта':'16',
}
#Спорт, фитнес, салоны красоты, SPA
xpath_sport ={

'Косметология':'2',
'Маникюр, педикюр':'4',
'Массаж':'8',
'Парикмахерское дело':'10',
'Салоны красоты / Парикмахерские / SPA':'14',
'Стилистика, визаж':'16',
'Тренерская работа, инструктаж':'20',
'Другое':'22',
'Начало карьеры, мало опыта':'26',

}
#Страхование
xpath_strahovanie ={

'Автострахование':'2',
'Актуарная деятельность':'4',
'Андеррайтинг':'8',
'Медицинское страхование':'10',
'Перестрахование':'14',
'Страхование бизнеса':'16',
'Страхование грузов':'20',
'Страхование жизни':'22',
'Страхование имущества':'26',
'Страхование коммерческих и финансовых рисков':'28',
'Страхование наружной рекламы':'32',
'Страхование недвижимости, ипотечное страхование':'34',
'Страхование ответственности':'38',
'Страхование строительно-монтажных работ':'40',
'Страхование туристов':'44',
'Страхование физических лиц':'46',
'Страхование юридических лиц':'50',
'Страховой анализ, оценка':'52',
'Урегулирование убытков':'56',
'Другое':'58',
'Начало карьеры, мало опыта':'62',

}
#Строительство, проектирование, недвижимость
xpath_stroitelstvo ={

'Агентства недвижимости, оценка недвижимости':'2',
'Архитектура':'4',
'Благоустройство территорий':'8',
'Водоснабжение, канализация':'10',
'Газоснабжение':'14',
'Геодезия, картография, землеустроительство':'16',
'ЖКХ':'20',
'Конструирование':'22',
'Отопление, вентиляция и кондиционирование':'26',
'Подводно-технические работы':'28',
'Проектирование':'32',
'Разработка генерального плана':'34',
'Слаботочные системы и сети':'38',
'Сметное дело':'40',
'Строительно-монтажные и отделочные работы':'44',
'Технический надзор, строительная экспертиза':'46',
'Транспортные системы (мосты, дороги, тоннели)':'50',
'Трубопроводы':'52',
'Управление недвижимостью':'56',
'Управление проектами (ГИП)':'58',
'Управление строительством, девелопмент':'62',
'Экологическое сопровождение проектов':'64',
'Эксплуатация зданий':'68',
'Электроснабжение':'70',
'Другое':'74',
'Начало карьеры, мало опыта':'76',

}
#Сырье
xpath_sirje ={

'Добыча':'4',
'Металлы':'8',
'Нефть':'10',
'Оборудование':'14',
'Переработка':'16',
'Разведка и разработка месторождений':'20',
'Технологии':'22',
'Уголь':'26',
'Другое':'28',
'Начало карьеры, мало опыта':'32',

}
#Топ-персонал
xpath_top_personal ={
'IT, Интернет, связь, телеком':'2',
'Административная работа, секретариат, АХО':'4',
'Банки, инвестиции, лизинг':'8',
'Безопасность, службы охраны':'10',
'Бухгалтерия, финансы, аудит':'14',
'Государственная служба, некоммерческие организации':'16',
'Дизайн':'20',
'Закупки, снабжение':'22',
'Искусство, культура, развлечения':'26',
'Кадры, управление персоналом':'28',
'Консалтинг, стратегическое развитие':'32',
'Маркетинг, реклама, PR':'34',
'Медицина, фармацевтика, ветеринария':'38',
'Наука, образование, повышение квалификации':'40',
'Продажи':'44',
'Промышленность, производство, сельское хозяйство':'46',
'СМИ, издательство, полиграфия':'50',
'Спорт, фитнес, салоны красоты, SPA':'52',
'Страхование':'56',
'Строительство, проектирование, недвижимость':'58',
'Сырье':'62',
'Транспорт, логистика, ВЭД':'64',
'Туризм, гостиницы, общественное питание':'68',
'Услуги, ремонт, сервисное обслуживание':'70',
'Юриспруденция':'74',
}
#Транспорт, логистика, ВЭД
xpath_transport ={

'Авиаперевозки':'2',
'Автоперевозки':'4',
'ВЭД':'8',
'Железнодорожные перевозки':'10',
'Контейнерные перевозки':'14',
'Логистика':'16',
'Метрополитен':'20',
'Морские, речные перевозки':'22',
'Складское хозяйство':'26',
'Таможня':'28',
'Трубопроводы':'32',
'Другое':'34',
'Начало карьеры, мало опыта':'38',

}
#Туризм, гостиницы, общественное питание
xpath_turizm ={

'HORECA':'2',
'Гостиницы, отели, кемпинги, мотели':'4',
'Кейтеринг':'8',
'Рестораны, кафе, столовые, фастфуд':'10',
'Туристические услуги и продукты':'14',
'Другое':'16',
'Начало карьеры, мало опыта':'20',
}
#Услуги, ремонт, сервисное обслуживание
xpath_remont ={

'Автосервис':'2',
'Ателье':'4',
'Бани':'8',
'Бытовая техника':'10',
'Клининг':'14',
'Компьютеры и оргтехника':'16',
'Медицинская техника':'20',
'Металлоремонт':'22',
'Оборудование для индустрии развлечения и спортивного инвентаря':'26',
'Обувь':'28',
'Промышленное оборудование':'32',
'Ритуальные услуги':'34',
'Телекоммуникационное оборудование и сети':'38',
'Торговое, складское, холодильное оборудование':'40',
'Фото, аудио, видео услуги':'44',
'Химчистки, прачечные':'46',
'Другое':'50',
'Начало карьеры, мало опыта':'52',
}
#Юриспруденция
xpath_urizm = {
'Compliance (нормативное право, соблюдение закона )':'2',
'Авторское право, патентное право':'4',
'Адвокатские услуги':'8',
'Антимонопольное право':'10',
'Арбитраж':'14',
'Банковское право':'16',
'Военное право':'20',
'Гражданское право':'22',
'Договорное право':'26',
'Законотворчество':'28',
'Земельное право':'32',
'Имущественное право':'34',
'Интеллектуальная собственность':'38',
'Корпоративное право':'40',
'Лицензирование':'44',
'Медицинское право':'46',
'Международное право':'50',
'Морское право':'52',
'Налоговое право':'56',
'Недропользование':'58',
'Нотариат':'62',
'Регистрация, перерегистрация, ликвидация предприятий':'64',
'Семейное право':'68',
'Слияния и поглощения':'70',
'Страховое право':'74',
'Таможенное право':'76',
'Трудовое право':'80',
'Уголовное право':'82',
'Урегулирование убытков':'86',
'Финансовое право':'88',
'Ценные бумаги':'92',
'Экологическое право':'94',
'Юридические консультации и услуги':'98',
'Другое':'100',
'Начало карьеры, мало опыта':'104',

}

vibor_podrubrik = {
    'IT, Интернет, связь, телеком' : xpath_it_ithernet,
    'Административная работа, секретариат, АХО' : xpath_it_adm_personal,
    'Банки, инвестиции, лизинг' : xpath_banki_invest,
    'Транспорт, логистика, ВЭД' : xpath_logistika,
    'Безопасность, службы охраны' : xpath_ohrana,
    'Бухгалтерия, финансы, аудит' : xparh_buhgalteria,
    'Дизайн' : xparh_design,
    'Домашний персонал' : xparh_home_personal,
    'Закупки, снабжение' : xparh_zakupki,
    'Исскуство, культура, развлечения' : xparh_kultura,
    'Кадры, управление персоналом' : xparh_hr_kadri,
    'Консалтинг, стратегическое развитие' : xparh_konsalting,
    'Маркетинг, реклама, PR' : xparh_marketing,
    'Медицина,  фармацевтика, ветеринария' : xparh_medicina,
    'Наука, образование, повышение квалификации' : xparh_obrazovanie,
    'Некоммерческие организации, волонтерство':xpath_nekom_org,
    'Продажи':xpath_prodazi,
    'Промышленность, производство':xpath_promishlen,
    'Рабочий персонал':xpath_rab_person,
    'СМИ, издательства':xpath_smi_izdatel,
    'Сельское хозяйство':xpath_selskoe,
    'Спорт, фитнес, салоны красоты, SPA':xpath_sport,
    'Страхование':xpath_strahovanie,
    'Строительство, проектирование, недвижимость':xpath_stroitelstvo,
    'Сырье':xpath_sirje,
    'Топ-персонал':xpath_top_personal,
    'Транспорт, логистика, ВЭД':xpath_transport,
    'Туризм, гостиницы, общественное питание':xpath_turizm,
    'Услуги, ремонт, сервисное обслуживание':xpath_remont,
    'Юриспруденция':xpath_urizm

    }

################### считываем данные с excel ################################

wb = xlrd.open_workbook('./DATA.xlsx')
sheet_xlr = wb.sheet_by_name('Вакансии')
a = 2


rows_sheet = sheet_xlr.nrows
# print(rows_sheet)


wb = load_workbook('./DATA.xlsx')
lst = (wb.sheetnames)

sheet = wb[lst[0]]
sheet.title
error_string = 1056

# random.randint(A, B) - случайное целое число N, A ≤ N ≤ B.

error_string = 1060
i = random.randint(3, rows_sheet)
################### заводим содержимое в переменные ################################
# print(a)
vacancy = (sheet.cell(row=i, column=1).value)
zp_ot = (sheet.cell(row=i, column=2).value)
zp_do = (sheet.cell(row=i, column=3).value)
opyt = (sheet.cell(row=i, column=4).value)
opisanie = (sheet.cell(row=i, column=5).value)
grafik = (sheet.cell(row=i, column=6).value)
company = (sheet.cell(row=i, column=7).value)
opis_company = (sheet.cell(row=i, column=8).value)
company1 = (sheet.cell(row=i, column=9).value)
opis_company1 = (sheet.cell(row=i, column=10).value)
company2 = (sheet.cell(row=i, column=11).value)
opis_company2 = (sheet.cell(row=i, column=12).value)
company3 = (sheet.cell(row=i, column=13).value)
opis_company3 = (sheet.cell(row=i, column=14).value)
company4 = (sheet.cell(row=i, column=15).value)
opis_company4 = (sheet.cell(row=i, column=16).value)
sfera_deyatel = (sheet.cell(row=i, column=17).value)
opis_deyatel = (sheet.cell(row=i, column=18).value)
name_company = (sheet.cell(row=i, column=19).value)
opis_deyatel_rename_company = (sheet.cell(row=i, column=20).value)
auto_podbor_rezume = (sheet.cell(row=1, column=20).value)

# print(vacancy)
error_string = 1099
# СЧИТЫВАЕМ ГОРОДА иЗ ЭКСЕЛЯ ДИНАМИЧЕСКИ

sheet = wb[lst[1]]
sheet.title

barnaul_adr = []
volgograd_adr = []
voronej_adr = []
ekaterenburg_adr = []
izevsk_adr = []
irkuts_adr = []
kazan_adr = []
kaliningrad_adr = []
kemerovo_adr = []
krasnodar_adr = []
krasnoyars_adr = []
moskow_adr = []
naber_cheln_adr = []
nizniy_novgorod_adr = []
novosib_adr = []
omsk_adr = []
orenburg_adr = []
perm_adr = []
rostov_na_dony_adr = []
ryazan_adr = []
samara_adr = []
sankt_peter_adr = []
saratov_adr = []
sochi_and_adler = []
toliatty_adr = []
tomsk_adr = []
tyla_adr = []
tumen_adr = []
ufa_adr = []
chelabinsk_adr = []
yaroslavl_adr = []
musor_exela = []
#goroda_arr = [musor_exela, barnaul_adr, volgograd_adr, voronej_adr, ekaterenburg_adr, izevsk_adr, irkuts_adr, kazan_adr, kaliningrad_adr, kemerovo_adr, krasnodar_adr, krasnoyars_adr, moskow_adr, naber_cheln_adr, nizniy_novgorod_adr, novosib_adr, omsk_adr, orenburg_adr, perm_adr, rostov_na_dony_adr, ryazan_adr, samara_adr, sankt_peter_adr, saratov_adr, sochi_and_adler, toliatty_adr, tomsk_adr, tyla_adr, tumen_adr, ufa_adr, chelabinsk_adr, yaroslavl_adr]
goroda_arr = []

kolichestvo_gorodov = 0

for cell in sheet['A']:
    kolichestvo_gorodov += 1
    print(cell.value)

for number_gorod in range(kolichestvo_gorodov):
    goroda_arr.append([])
    for cell in list(sheet.rows)[number_gorod]:
         if cell.value != None:
            goroda_arr[number_gorod].append(cell.value)

kolichestvo_adresov = []
high_number_gorod = 0
number_goroda = 0
while number_goroda < len(goroda_arr):
    max_number = len(goroda_arr[number_goroda])
    kolichestvo_adresov.append(len(goroda_arr[number_goroda]))
    if high_number_gorod< max_number:
        high_number_gorod = max_number
        number_goroda = number_goroda + 1
    else:
        number_goroda = number_goroda + 1

###############    autorith  ###############

n = 0
while n < 150:
  try:
    sheet = wb[lst[2]]
    sheet.title
    login = (sheet.cell(row=1, column=2).value)
    password = (sheet.cell(row=2, column=2).value)

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    browser = webdriver.Chrome(chrome_options=options)
    driver = browser

    browser.get('https://nn.superjob.ru/hr/vacancy/create/')



    ############### Login ########################

    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[3]/div/div/div/div/div[1]/label/div/div/input').send_keys(login)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[3]/div/div/div/div/div[2]/label/div/div[1]/input').send_keys(password)
    time.sleep(1)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/form/div/div[4]/div/div[1]/button').click()

    ############### Должность ####################
    time.sleep(4)
    browser.find_element_by_xpath('//*[@id="app"]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[1]/div/div[2]/div/div[1]/label/div/div/input').send_keys(vacancy)

    ############## Специализация #################
    companys = [company, company1, company2, company3, company4]
    opis_companys = [opis_company, opis_company1, opis_company2, opis_company3, opis_company4]
    iter_companys = 0  # Для теста пока без цикла, потом добавить как в работару
    print(companys)
    print(opis_companys)

    browser.find_element_by_xpath(
        '/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[2]/div/div/div/button').click()
    try:
        browser.find_element_by_xpath(
            '/html/body/div[3]/div/div[2]/div/div[2]/div/div/div/div[2]/div/div[3]/div/button/span/span/span').click()
    except:
        print('чисто')
    time.sleep(1)
    buffer_specializazii = 999

    while iter_companys < 5:
        if iter_companys != 0:
            browser.find_element_by_xpath(
                '/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[2]/div/div[2]/div/div/div/button').click()
        # Тыкаем на специализацию!
        if opis_companys[iter_companys] == "СМИ, издательства":
            if companys[iter_companys] == "Радио":
                specializacii = browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                    opis_companys[iter_companys])
                browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[1]/div[2]/div/div/div/label/div/div[2]/span').click()
        elif opis_companys[iter_companys] == "Топ-персонал":
            if companys[iter_companys] == "Страхование":
                specializacii = browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                    opis_companys[iter_companys])
                browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[2]/div[2]/div/div/div/label/div/div[2]/span').click()

        elif opis_companys[iter_companys] == "Транспорт, логистика, ВЭД":
            if companys[iter_companys] == "Трубопроводы":
                specializacii = browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                    opis_companys[iter_companys])
                browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[1]/div[2]/div/div/div/label/div/div[2]/span').click()
        elif opis_companys[iter_companys] == "Строительство, проектирование, недвижимость":
            if companys[iter_companys] == "Трубопроводы":
                specializacii = browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                    opis_companys[iter_companys])
                browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[2]/div[2]/div/div/div/label/div/div[2]/span').click()

        elif opis_companys[iter_companys] == "Юриспруденция":
            if companys[iter_companys] == "Урегулирование убытков":
                specializacii = browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[1]/div[2]/div/div/div/label/div/div[2]/span').send_keys(
                    opis_companys[iter_companys])
                browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[1]/div[2]/div/div/div/label/div/div[2]/span').click()

        elif opis_companys[iter_companys] == "Страхование":
            if companys[iter_companys] == "Урегулирование убытков":
                specializacii = browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                    opis_companys[iter_companys])
                browser.find_element_by_xpath(
                    '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div[2]/div[2]/div/div/div/label/div/div[2]/span').click()
        elif opis_companys[iter_companys] == "Начало карьеры, мало опыта":
            specializacii = browser.find_element_by_xpath(
                '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                'Начало карьеры')
            oblast = browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div')
            specializacii = oblast.find_elements_by_class_name('_2s93E')
            n_spec = xpath_specializacii.get(companys[iter_companys])
            specializacii[int(n_spec)].click()
        elif opis_companys[iter_companys] == "Другое":
            specializacii = browser.find_element_by_xpath(
                '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                opis_companys[iter_companys])
            oblast = browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div')
            specializacii = oblast.find_elements_by_class_name('_2s93E')
            n_spec = xpath_specializacii.get(companys[iter_companys])
            specializacii[int(n_spec)].click()
        else:
            specializacii = browser.find_element_by_xpath(
                '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[1]/div/label/div/div/input').send_keys(
                opis_companys[iter_companys])
            browser.find_element_by_xpath(
                '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]/div/div[2]/div/div/div/label/div/div[1]/span').click()
        time.sleep(2)
        browser.find_element_by_xpath(
            '/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[3]/div/div/div[1]/button/span/span/span').click()
        time.sleep(1)

        iter_companys += 1

    ############# Занятость #################### НУЖНО СДЕЛАТЬ

    browser.find_element_by_xpath('//*[@id="detailInfo.workType.id-input"]').click()
    browser.find_element_by_xpath('//*[@id="detailInfo.workType.id-input"]').send_keys(grafik + Keys.ENTER)
    ############# Уровень дохода ###############

    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[7]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[1]/label/div/div[2]/input').send_keys(zp_ot)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[7]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div/div[2]/label/div/div[2]/input').send_keys(zp_do)

    ############ Опыт работы ###################
    try:
        if opyt == 'не имеет значения':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[1]').click()
        elif opyt == '1 год':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[3]').click()
        elif opyt == '3 года':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[4]').click()
        elif opyt == '6 лет':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[5]').click()
        elif opyt == 'без опыта':
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[8]/div/div[2]/button[2]').click()
    except:
        print('Выбраный опыт работы уже стоит')

    ############ Описание вакансии #############

    #Вытаскиваем Условия
    frag_opisanie = opisanie.partition('Условия:')
    frag_opisanie = frag_opisanie[2].partition('Обязанности:')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    yslovia = frag_opisanie[0]

    #Вытаскиваем Обязанности
    frag_opisanie = opisanie.partition('Обязанности:')
    frag_opisanie = frag_opisanie[2].partition('Требования:')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    obiaznosti = frag_opisanie[0]

    #Вытаскиваем Требования
    frag_opisanie = opisanie.partition('Требования:')
    frag_opisanie = frag_opisanie[2].partition('\n \n')
    frag_opisanie = frag_opisanie[0].split('\n', 1)
    del frag_opisanie[0]
    trebovania = frag_opisanie[0]

    #Вытаскиваем шапку
    frag_opisanie = opisanie.partition('Условия:')
    shapka = frag_opisanie[0]
    print(frag_opisanie)

    #Вытаскиваем подвал
    frag_opisanie = opisanie.partition('\n \n')
    podval = frag_opisanie[2]

    #Заполняем описание
    linii = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[10]/div/div/div[2]/div/div[1]/div/div/div/div[2]')
    linii.find_element_by_tag_name('b').send_keys(Keys.ARROW_UP + Keys.ENTER)
    lst_linii = linii.find_elements_by_tag_name('li')
    linii.find_element_by_tag_name('p').send_keys(Keys.LEFT_CONTROL + 'b' + Keys.LEFT_CONTROL + (shapka + Keys.BACKSPACE))
    lst_linii[0].send_keys(obiaznosti + Keys.BACKSPACE + Keys.DELETE)
    lst_linii[2].send_keys(trebovania + Keys.DELETE)
    lst_linii[4].send_keys(yslovia + Keys.BACKSPACE + Keys.DELETE)
    linii.send_keys('\n\n' + Keys.LEFT_CONTROL + 'b' + Keys.LEFT_CONTROL + podval)



    #Заполняем города
    #ля теста iter_gorod
    iter_gorod = random.randint(1, len(goroda_arr))
    print(len(goroda_arr))
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[1]/div/div/div[2]/ul/li[1]/span/span[2]/button').click()
    except:
        print('город автоматически не проставился')

    adr_bar = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[1]/div/div/div[3]/div/div[1]/label/div/div/input')
    random_adres = random.randint(1, kolichestvo_adresov[iter_gorod]-1)
    print('random adres = ', random_adres)
    adr_bar.send_keys(goroda_arr[iter_gorod][0])
    print(goroda_arr[iter_gorod][0])

    adr_bar = browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[4]/div/div[2]/div/div/div/div[2]/div/div[1]/div/div[1]/label/div/div/input')
    adr_bar.send_keys(goroda_arr[iter_gorod][random_adres])
    time.sleep(2)
    adr_bar.send_keys(Keys.ARROW_DOWN + Keys.ENTER)

    #########Добавляем метро
    time.sleep(2)
    try:
        browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[1]/div/div/div[5]/div/div/div/button/span/span/span').click()
        time.sleep(2)
        list_metro = browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[2]')
        spisoc_metro = list_metro.find_elements_by_class_name('_2s93E')

        metro = len(spisoc_metro) - 1
        array_metro = list(range(0, metro))
        random.shuffle(array_metro)
        iter_metro = 0
        while iter_metro < 10 + iter_metro < metro:
            iter_iter_metro = array_metro[iter_metro]
            try:
                spisoc_metro[iter_iter_metro].click()
                time.sleep(1)
                iter_metro += 1
            except:
                time.sleep(1)
                iter_metro+=1

        browser.find_element_by_xpath('/html/body/div[3]/div/div[5]/div[2]/div/div[2]/div/div/div[3]/div/button/span/span/span').click()
    except:
        print('в городе нет метро')
    ############## Информация о компании ###################################

    if sfera_deyatel != None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/label/div/div[1]/span').click()
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[2]/button').click()
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[1]/div/label/div/div/input').send_keys(sfera_deyatel)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[2]/div/label/div/div/textarea').send_keys(opis_deyatel)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[3]/div/div[3]/div/div[1]/button').click()
        except:
            print("Скрыть информацию не получилось")
    elif name_company != None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[2]/button').click()
            time.sleep(1)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[1]/div/label/div/div/input').send_keys(name_company)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[2]/div/label/div/div/textarea').send_keys(opis_deyatel_rename_company)
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[3]/div/div/div[2]/div/div[1]/div/div[3]/div/div[1]/button').click()
        except:
            print("Отредактировать информацию не получилось")
    else:
        print("Оставляем название компании ткаим какое оно есть")

    #####Убераем бонус
    if auto_podbor_rezume == None:
        try:
            browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[5]/div/div[1]/div/div[1]/div/label/div/div[1]/span').click()
        except:
            print('бонуса нет')
    ############## Размещаем вакансию и закрываем браузер ##################
    time.sleep(1)
    browser.find_element_by_xpath('/html/body/div[3]/div/div[1]/div[4]/div/div/div/div/div/form/div/div[6]/div/div[2]/div/div[1]/div/div[1]/button/span/span/span').click()
    time.sleep(1)
    print('опубликовано')
    browser.quit()
    n+=1
  except:
      browser.quit()
      n+=1

#browser.quit()